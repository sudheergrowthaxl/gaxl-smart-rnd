"""OutputFormatterAgent - Exports DQ rules to JSON and Excel formats."""

import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.dq_rule import DQRule, DQRuleSet
from ..models.agent_state import ValidationResult


class OutputFormatterAgent:
    """
    Agent responsible for formatting and exporting DQ rules.

    This agent:
    - Exports rules to JSON format with full schema
    - Exports rules to Excel with multiple sheets and formatting
    - Generates summary statistics
    """

    # Style definitions
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    SEVERITY_COLORS = {
        "Critical": "FF6B6B",  # Red
        "High": "FFA94D",      # Orange
        "Medium": "FFE066",    # Yellow
        "Low": "8CE99A",       # Green
    }

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def __init__(self, output_dir: str = "output"):
        """
        Initialize the OutputFormatterAgent.

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_json(
        self,
        rules: List[DQRule],
        dataset_context: Dict[str, Any],
        output_filename: str = "dq_rules.json",
    ) -> str:
        """
        Export rules to JSON format.

        Args:
            rules: List of DQRule objects
            dataset_context: Dataset metadata
            output_filename: Name of output file

        Returns:
            Path to the generated JSON file
        """
        ruleset = DQRuleSet(
            dataset_name=dataset_context.get('dataset_name', 'Product_Data'),
            parent_class=dataset_context.get('parent_class', 'Unknown'),
            total_records=dataset_context.get('total_records', 0),
            generated_at=datetime.now().isoformat(),
            source_profiling=dataset_context.get('profiling_path', ''),
            rules=rules,
        )
        ruleset.generate_summary()

        output_path = self.output_dir / output_filename

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(ruleset.to_dict(), f, indent=2, default=str)

        return str(output_path)

    def export_to_excel(
        self,
        rules: List[DQRule],
        validation_results: Optional[List[ValidationResult]] = None,
        output_filename: str = "dq_rules.xlsx",
        dataset_context: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Export rules to Excel format with formatting.

        Args:
            rules: List of DQRule objects
            validation_results: Optional list of ValidationResult objects
            output_filename: Name of output file
            dataset_context: Optional dataset context with parent_class info

        Returns:
            Path to the generated Excel file
        """
        output_path = self.output_dir / output_filename
        parent_class = dataset_context.get('parent_class', 'Unknown') if dataset_context else 'Unknown'

        wb = Workbook()

        # Sheet 1: DQ Rules (main table) - includes Parent Class/Category column
        ws_rules = wb.active
        ws_rules.title = "DQ Rules"
        self._write_rules_sheet(ws_rules, rules, parent_class)

        # Sheet 2: Summary by Category
        ws_summary = wb.create_sheet("Summary by Category")
        self._write_summary_sheet(ws_summary, rules, parent_class)

        # Sheet 3: Validation Results
        if validation_results:
            ws_validation = wb.create_sheet("Validation Results")
            self._write_validation_sheet(ws_validation, validation_results)

        # Sheet 4: SQL Implementations
        ws_sql = wb.create_sheet("SQL Implementations")
        self._write_sql_sheet(ws_sql, rules, parent_class)

        # Sheet 5: Python Implementations
        ws_python = wb.create_sheet("Python Implementations")
        self._write_python_sheet(ws_python, rules, parent_class)

        # Sheet 6: Rules by Attribute
        ws_by_attr = wb.create_sheet("Rules by Attribute")
        self._write_by_attribute_sheet(ws_by_attr, rules, parent_class)

        wb.save(output_path)
        return str(output_path)

    def _write_rules_sheet(self, ws, rules: List[DQRule], parent_class: str = "Unknown") -> None:
        """Write main rules table to worksheet with Parent Class/Category column."""
        headers = [
            "Parent Class/Category", "Rule ID", "Attribute", "Category", "Rule Type", "Expression",
            "Severity", "Description", "Threshold %", "Confidence",
            "Derived From", "Valid Values", "Invalid Values"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self.THIN_BORDER

        # Write data rows
        for row, rule in enumerate(rules, 2):
            ws.cell(row=row, column=1, value=parent_class)  # Parent Class/Category
            ws.cell(row=row, column=2, value=rule.rule_id)
            ws.cell(row=row, column=3, value=rule.attribute_name)
            ws.cell(row=row, column=4, value=rule.rule_category)
            ws.cell(row=row, column=5, value=rule.rule_type)
            ws.cell(row=row, column=6, value=rule.rule_expression)

            # Severity with color
            severity_cell = ws.cell(row=row, column=7, value=rule.severity)
            color = self.SEVERITY_COLORS.get(rule.severity, "FFFFFF")
            severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            ws.cell(row=row, column=8, value=rule.description)
            ws.cell(row=row, column=9, value=rule.threshold_percent)
            ws.cell(row=row, column=10, value=rule.confidence_score)
            ws.cell(row=row, column=11, value=rule.derived_from)
            ws.cell(row=row, column=12, value=", ".join(rule.sample_valid_values[:5]))
            ws.cell(row=row, column=13, value=", ".join(rule.sample_invalid_values[:5]))

            # Apply borders
            for col in range(1, 14):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

        # Adjust column widths
        column_widths = {
            'A': 25, 'B': 45, 'C': 30, 'D': 15, 'E': 18, 'F': 60,
            'G': 12, 'H': 50, 'I': 12, 'J': 12, 'K': 40,
            'L': 30, 'M': 30
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _write_summary_sheet(self, ws, rules: List[DQRule], parent_class: str = "Unknown") -> None:
        """Write summary statistics by category."""
        categories = ["Completeness", "Validity", "Accuracy",
                      "Consistency", "Uniqueness", "Timeliness"]
        severities = ["Critical", "High", "Medium", "Low"]

        # Build summary data
        summary = {}
        for cat in categories:
            summary[cat] = {sev: 0 for sev in severities}
            summary[cat]["Total"] = 0

        for rule in rules:
            cat = rule.rule_category
            sev = rule.severity
            if cat in summary and sev in summary[cat]:
                summary[cat][sev] += 1
                summary[cat]["Total"] += 1

        # Write headers
        headers = ["Category"] + severities + ["Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        # Write data
        for row, (cat, counts) in enumerate(summary.items(), 2):
            ws.cell(row=row, column=1, value=cat).border = self.THIN_BORDER
            for col, sev in enumerate(severities, 2):
                cell = ws.cell(row=row, column=col, value=counts[sev])
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=6, value=counts["Total"]).border = self.THIN_BORDER

        # Add totals row
        total_row = len(summary) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col, sev in enumerate(severities, 2):
            total = sum(summary[cat][sev] for cat in categories)
            ws.cell(row=total_row, column=col, value=total).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=len(rules)).font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_validation_sheet(self, ws, validation_results: List[ValidationResult]) -> None:
        """Write validation results."""
        headers = ["Rule ID", "Pass Count", "Fail Count", "Pass Rate %", "Sample Failures"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        for row, result in enumerate(validation_results, 2):
            ws.cell(row=row, column=1, value=result.rule_id)
            ws.cell(row=row, column=2, value=result.pass_count)
            ws.cell(row=row, column=3, value=result.fail_count)

            # Color-code pass rate
            pass_cell = ws.cell(row=row, column=4, value=result.pass_rate)
            if result.pass_rate >= 95:
                pass_cell.fill = PatternFill(start_color="8CE99A", end_color="8CE99A", fill_type="solid")
            elif result.pass_rate >= 80:
                pass_cell.fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
            else:
                pass_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

            # Truncate sample failures for readability
            failures_str = str(result.sample_failures)[:200]
            ws.cell(row=row, column=5, value=failures_str)

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 80

    def _write_sql_sheet(self, ws, rules: List[DQRule], parent_class: str = "Unknown") -> None:
        """Write SQL implementations with Parent Class/Category column."""
        headers = ["Parent Class/Category", "Rule ID", "Attribute", "Category", "SQL Expression"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        for row, rule in enumerate(rules, 2):
            ws.cell(row=row, column=1, value=parent_class)
            ws.cell(row=row, column=2, value=rule.rule_id)
            ws.cell(row=row, column=3, value=rule.attribute_name)
            ws.cell(row=row, column=4, value=rule.rule_category)
            ws.cell(row=row, column=5, value=rule.rule_expression_sql)

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 120

    def _write_python_sheet(self, ws, rules: List[DQRule], parent_class: str = "Unknown") -> None:
        """Write Python implementations with Parent Class/Category column."""
        headers = ["Parent Class/Category", "Rule ID", "Attribute", "Category", "Python Expression"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        for row, rule in enumerate(rules, 2):
            ws.cell(row=row, column=1, value=parent_class)
            ws.cell(row=row, column=2, value=rule.rule_id)
            ws.cell(row=row, column=3, value=rule.attribute_name)
            ws.cell(row=row, column=4, value=rule.rule_category)
            ws.cell(row=row, column=5, value=rule.rule_expression_python)

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 120

    def _write_by_attribute_sheet(self, ws, rules: List[DQRule], parent_class: str = "Unknown") -> None:
        """Write rules grouped by attribute with Parent Class/Category column."""
        # Group rules by attribute
        by_attr: Dict[str, List[DQRule]] = {}
        for rule in rules:
            attr = rule.attribute_name
            if attr not in by_attr:
                by_attr[attr] = []
            by_attr[attr].append(rule)

        headers = ["Parent Class/Category", "Attribute", "Rule Count", "Categories", "Severities", "Rule IDs"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        for row, (attr, attr_rules) in enumerate(sorted(by_attr.items()), 2):
            categories = list(set(r.rule_category for r in attr_rules))
            severities = list(set(r.severity for r in attr_rules))
            rule_ids = [r.rule_id for r in attr_rules]

            ws.cell(row=row, column=1, value=parent_class)
            ws.cell(row=row, column=2, value=attr)
            ws.cell(row=row, column=3, value=len(attr_rules))
            ws.cell(row=row, column=4, value=", ".join(categories))
            ws.cell(row=row, column=5, value=", ".join(severities))
            ws.cell(row=row, column=6, value=", ".join(rule_ids))

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 80

    def generate_summary(self, rules: List[DQRule]) -> Dict[str, Any]:
        """
        Generate summary statistics for the ruleset.

        Args:
            rules: List of DQRule objects

        Returns:
            Summary dictionary
        """
        categories = ["Completeness", "Validity", "Accuracy",
                      "Consistency", "Uniqueness", "Timeliness"]
        severities = ["Critical", "High", "Medium", "Low"]

        return {
            "total_rules": len(rules),
            "rules_by_category": {
                cat: len([r for r in rules if r.rule_category == cat])
                for cat in categories
            },
            "rules_by_severity": {
                sev: len([r for r in rules if r.severity == sev])
                for sev in severities
            },
            "attributes_covered": list(set(r.attribute_name for r in rules)),
            "attribute_count": len(set(r.attribute_name for r in rules)),
            "avg_confidence_score": (
                sum(r.confidence_score for r in rules) / len(rules)
                if rules else 0
            ),
            "avg_threshold": (
                sum(r.threshold_percent for r in rules) / len(rules)
                if rules else 0
            ),
        }
